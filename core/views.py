import os
import openpyxl
import arabic_reshaper
from bidi.algorithm import get_display
from io import BytesIO
from openpyxl.styles import Font, Alignment, Border, Side
import openpyxl.styles
from openpyxl.utils import get_column_letter
from django.utils import timezone
from datetime import datetime, date
from decimal import Decimal

from fpdf import FPDF

from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, Paragraph

from django.conf import settings
from django.http import HttpResponse, HttpResponseForbidden
from django.utils.text import slugify
from django.urls import reverse_lazy
from django.forms import HiddenInput
from django.views import View
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.db.models import Sum, Count, Value

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth import update_session_auth_hash, logout
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.models import User
from django.contrib import messages

from .utils import get_loan_chart_data, get_covenant_chart_data, get_budget_revenue_chart_data
from .models import Loan, Covenant, Client, Budget, BudgetExpense, BudgetRevenue, Fund, FundExpense, FundRevenue
from .models import ClientType, ClientCategory, LoanType, CovenantType, ExpenseCategory, RevenueCategory
from .forms import (
    LoanForm, CovenantForm, ClientForm, BudgetForm, BudgetRevenueForm, BudgetExpenseForm,
    FundForm, FundExpenseForm, FundRevenueForm, UserForm, ProfileForm,
    ClientPhoneFormSet, ClientEmailFormSet, ClientDocumentFormSet,
    LoanFilterForm, CovenantFilterForm, ClientFilterForm,
    FundFilterForm, FundExpenseFilterForm, FundRevenueFilterForm,
    BudgetFilterForm, BudgetExpenseFilterForm, BudgetRevenueFilterForm,
    ClientTypeForm, ClientCategoryForm, LoanTypeForm, CovenantTypeForm, ExpenseCategoryForm, RevenueCategoryForm
)

def custom_403(request, exception):
    return render(request, '403.html', {}, status=403)

def custom_404(request, exception):
    return render(request, '404.html', {}, status=404)

def custom_500(request):
    return render(request, '500.html', status=500)

@login_required
def index(request):
    total_loans = Loan.objects.aggregate(
        total_amount=Sum('amount', default=Value(0)),
        total_paid_amount=Sum('paid_amount', default=Value(0)),
        total_remaining_amount=Sum('remaining_amount', default=Value(0)),
    )
    loan_chart_data = get_loan_chart_data()

    total_covenants = Covenant.objects.aggregate(
        total_amount=Sum('amount', default=Value(0)),
        total_paid_amount=Sum('paid_amount', default=Value(0)),
        total_remaining_amount=Sum('remaining_amount', default=Value(0)),
    )
    covenant_chart_data = get_covenant_chart_data()

    funds = Fund.objects.aggregate(
        total=Sum('current_balance', default=Value(0)),
        count=Count('id'),
    )
    clients = Client.objects.aggregate(
        total=Sum('balance', default=Value(0)),
        count=Count('id'),
    )

    budget_revenue_chart_data = get_budget_revenue_chart_data()
    budgetrevenues = BudgetRevenue.objects.aggregate(
        total_actual=Sum('actual_amount', default=Value(0)),
        total_estimated=Sum('estimated_amount', default=Value(0)),
    )
    budgetexpenses = BudgetExpense.objects.aggregate(
        total_actual=Sum('actual_amount', default=Value(0)),
        total_estimated=Sum('estimated_amount', default=Value(0)),
    )
    budgets = Budget.objects.aggregate(
        count=Count('name'),
    )

    context = {
        'amount_loans': "{:,.2f}".format(total_loans['total_amount']),
        'paid_amount_loans': "{:,.2f}".format(total_loans['total_paid_amount']),
        'remaining_amount_loans': "{:,.2f}".format(total_loans['total_remaining_amount']),
        'loan_by_type': loan_chart_data['by_type'],
        'loan_by_owner': loan_chart_data['by_owner'],

        'amount_covenants': "{:,.2f}".format(total_covenants['total_amount']),
        'paid_amount_covenants': "{:,.2f}".format(total_covenants['total_paid_amount']),
        'remaining_amount_covenants': "{:,.2f}".format(total_covenants['total_remaining_amount']),
        'covenant_by_type': covenant_chart_data['by_type'],
        'covenant_by_owner': covenant_chart_data['by_owner'],

        'funds_total': "{:,.2f}".format(funds['total']),
        'funds_count': funds['count'],

        'clients_total': "{:,.2f}".format(clients['total']),
        'clients_count': clients['count'],

        'budget_revenue_chart_data': budget_revenue_chart_data,
        'budgets_count': budgets['count'],
        'budgets_total': "{:,.2f}".format((budgetrevenues['total_actual'] or 0) - (budgetexpenses['total_actual'] or 0)),
        'budgetrevenues_actual': "{:,.2f}".format(budgetrevenues['total_actual']),
        'budgetrevenues_estimated': "{:,.2f}".format(budgetrevenues['total_estimated']),
        'budgetexpenses_actual': "{:,.2f}".format(budgetexpenses['total_actual']),
        'budgetexpenses_estimated': "{:,.2f}".format(budgetexpenses['total_estimated']),
    }
    return render(request, 'index.html', context=context)

@login_required
def logout_view(request):
    logout(request)
    messages.success(request, 'تم تسجيل الخروج بنجاح!')
    return redirect('login')

# View for displaying and editing a user's profile
@login_required
def profile_view(request):
    user = request.user  # Get the currently logged-in user

    if request.method == 'POST':
        # If POST request, handle user form submission
        form = ProfileForm(request.POST, instance=user)
        password_form = PasswordChangeForm(request.user, request.POST)  # Password change form
        if form.is_valid():
            user = form.save(commit=False)
            user.save()  # Save the updated profile details
            # Display success message if any updates occurred
            messages.success(request, 'تم تحديث الملف الشخصي بنجاح!')
            return redirect('profile')

            # Handle password change
        if password_form.is_valid() and password_form.has_changed():
            user = password_form.save()  # Save the password
            # Update session to keep the user logged in after password change
            update_session_auth_hash(request, user)

            # Display success message if any updates occurred
            messages.success(request, 'تم تحديث الملف الشخصي بنجاح!')
            return redirect('profile')
    else:
        # If GET request, initialize the forms with the current user's data
        form = ProfileForm(instance=user)
        password_form = PasswordChangeForm(request.user)

    return render(request, 'users/profile.html', {
        'form': form,  # User information form
        'password_form': password_form,  # Password change form
        'user': user
    })

# User Views
@login_required
@permission_required('auth.view_user', raise_exception=True)
def user_list(request):
    users = User.objects.all()
    return render(request, 'users/user_list.html', {'users': users})

@login_required
@permission_required('auth.add_user', raise_exception=True)
def user_create(request):
    if request.method == 'POST':
        form = UserForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('user_list')
    else:
        form = UserForm()
    return render(request, 'users/user_form.html', {'form': form})

@login_required
@permission_required('auth.change_user', raise_exception=True)
def user_update(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        form = UserForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            return redirect('user_list')
    else:
        form = UserForm(instance=user)
    return render(request, 'users/user_form.html', {'form': form, 'user_update': user})

@login_required
@permission_required('auth.delete_user', raise_exception=True)
def user_delete(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        user.delete()
        return redirect('user_list')
    return render(request, 'users/delete_user.html', {'user': user})


# Settings Views
@login_required
@permission_required('auth.change_user', raise_exception=True)
def settings_view(request):
    if not request.user.is_superuser:
        return HttpResponseForbidden("You do not have permission to view this page.")

    # Initialize forms without data
    client_type_form = ClientTypeForm()
    client_category_form = ClientCategoryForm()
    loan_type_form = LoanTypeForm()
    covenant_type_form = CovenantTypeForm()
    expense_category_form = ExpenseCategoryForm()
    revenue_category_form = RevenueCategoryForm()

    # Get current records
    client_types = ClientType.objects.all()
    client_categories = ClientCategory.objects.all()
    loan_types = LoanType.objects.all()
    covenant_types = CovenantType.objects.all()
    expense_categories = ExpenseCategory.objects.all()
    revenue_categories = RevenueCategory.objects.all()

    if request.method == "POST":
        form_type = request.POST.get("form_type")

        if form_type == "ClientType":
            client_type_form = ClientTypeForm(request.POST)
            if client_type_form.is_valid():
                client_type_form.save()
                messages.success(request, "تم إضافة نوع العميل بنجاح.")
        
        elif form_type == "ClientCategory":
            client_category_form = ClientCategoryForm(request.POST)
            if client_category_form.is_valid():
                client_category_form.save()
                messages.success(request, "تم إضافة فئة العميل بنجاح.")

        elif form_type == "LoanType":
            loan_type_form = LoanTypeForm(request.POST)
            if loan_type_form.is_valid():
                loan_type_form.save()
                messages.success(request, "تم إضافة شكل السلف بنجاح.")

        elif form_type == "CovenantType":
            covenant_type_form = CovenantTypeForm(request.POST)
            if covenant_type_form.is_valid():
                covenant_type_form.save()
                messages.success(request, "تم إضافة شكل العهد بنجاح.")

        elif form_type == "ExpenseCategory":
            expense_category_form = ExpenseCategoryForm(request.POST)
            if expense_category_form.is_valid():
                expense_category_form.save()
                messages.success(request, "تم إضافة فئة المصروف بنجاح.")

        elif form_type == "RevenueCategory":
            revenue_category_form = RevenueCategoryForm(request.POST)
            if revenue_category_form.is_valid():
                revenue_category_form.save()
                messages.success(request, "تم إضافة فئة الإيراد بنجاح.")

        # Handle deleting entries
        elif "delete" in request.POST:
            model_name = request.POST.get("model_name")
            record_id = request.POST.get("record_id")
            if model_name == "ClientType":
                ClientType.objects.filter(id=record_id).delete()
                messages.success(request, "تم حذف نوع العميل بنجاح.")
            elif model_name == "ClientCategory":
                ClientCategory.objects.filter(id=record_id).delete()
                messages.success(request, "تم حذف فئة العميل بنجاح.")
            elif model_name == "LoanType":
                LoanType.objects.filter(id=record_id).delete()
                messages.success(request, "تم حذف شكل السلف بنجاح.")
            elif model_name == "CovenantType":
                CovenantType.objects.filter(id=record_id).delete()
                messages.success(request, "تم حذف شكل العهد بنجاح.")
            elif model_name == "ExpenseCategory":
                ExpenseCategory.objects.filter(id=record_id).delete()
                messages.success(request, "تم حذف فئة المصروف بنجاح.")
            elif model_name == "RevenueCategory":
                RevenueCategory.objects.filter(id=record_id).delete()
                messages.success(request, "تم حذف فئة الإيراد بنجاح.")

        return redirect("settings")  # Refresh page after adding

    # Render the settings template
    context = {
        "client_type_form": client_type_form,
        "client_category_form": client_category_form,
        "loan_type_form": loan_type_form,
        "covenant_type_form": covenant_type_form,
        "expense_category_form": expense_category_form,
        "revenue_category_form": revenue_category_form,
        "client_types": client_types,
        "client_categories": client_categories,
        "loan_types": loan_types,
        "covenant_types": covenant_types,
        "expense_categories": expense_categories,
        "revenue_categories": revenue_categories,
    }

    return render(request, "users/settings.html", context)


# Permission Required
class PermissionMixin(PermissionRequiredMixin):
    def has_permission(self):
        perms = self.get_permission_required()
        return self.request.user.has_perms(perms) or self.request.user.is_superuser


# Export Excels & PDFs
class ExportMixin:
    """Mixin to export any ListView data to an Excel or PDF file with proper Arabic support."""
    def get(self, request, *args, **kwargs):
        """Export data when 'export' is in request GET parameters."""
        if "export" in request.GET:
            export_format = request.GET.get("format", "excel")  # Default to 'excel'
            queryset = self.get_queryset()

            if not queryset.exists():
                return HttpResponse("No data available for export.", content_type="text/plain")

            if export_format == "pdf":
                return self.export_to_pdf(queryset)
            else:
                return self.export_to_excel(queryset)

        return super().get(request, *args, **kwargs)

    def prepare_export_response(self, content, file_type, model_name):
        """Prepare and return a file response with a timestamped filename."""
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        safe_model_name = slugify(model_name) or "export"
        filename = f"export_{safe_model_name}_{timestamp}.{file_type}"

        content_types = {
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "pdf": "application/pdf",
        }

        response = HttpResponse(content, content_type=content_types.get(file_type, "application/octet-stream"))
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def export_to_excel(self, queryset):
        """Generate an Excel file from a queryset with all fields, modern style."""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"{queryset.model._meta.verbose_name_plural}"

        # Get field names dynamically
        fields = [field.name for field in queryset.model._meta.fields[1:]]
        fields1 = [field.verbose_name for field in queryset.model._meta.fields[1:]]

        # Set modern header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers with styles
        for col_num, field_name in enumerate(fields1, 1):
            cell = sheet.cell(row=1, column=col_num, value=field_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Set the column width dynamically based on the length of the content
        for col_num, field_name in enumerate(fields1, 1):
            column_width = max(len(field_name), 15)  # Ensure a minimum width of 15
            sheet.column_dimensions[get_column_letter(col_num)].width = column_width

        # Write rows
        for obj in queryset:
            row = []
            for field in fields:
                value = getattr(obj, field, "")

                # Handle related fields (ForeignKey, ManyToMany, etc.)
                if isinstance(value, str):
                    row.append(value)
                elif hasattr(value, 'get_FOO_display'):  # For choices-based fields
                    row.append(value.get_FOO_display())
                else:
                    # Check if the field is a related object (e.g., ForeignKey)
                    related_object = getattr(obj, field, None)
                    if related_object:
                        row.append(str(related_object))  # Get the string representation
                    else:
                        row.append("")

            sheet.append(row)

        # Apply border style for all cells
        border_style = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = border_style

        # Save workbook to memory
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        content = output.read()
        file_type = "xlsx"
        model_name = queryset.model._meta.model_name
        return self.prepare_export_response(content, file_type, model_name)

    def export_to_pdf(self, queryset):
        """Generate a properly formatted PDF with RTL Arabic support and custom font."""
        model_name = queryset.model._meta.model_name

        # Create an in-memory buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=10, leftMargin=10, topMargin=20, bottomMargin=20)


        # Load Arabic font
        font_path = os.path.abspath(os.path.join(settings.STATIC_ROOT, "fonts", "Janna LT Bold", "Janna LT Bold.ttf"))
        pdfmetrics.registerFont(TTFont("Janna", font_path))

        elements = []
        # Table Headers (Right-aligned for Arabic)
        fields = [[field.name, field.verbose_name] for field in queryset.model._meta.fields[-1:0:-1]]
        headers = [get_display(arabic_reshaper.reshape(field[1])) for field in fields]
        table_data = [headers]  # Table header

        max_col_lengths = [len(header) for header in headers]
        # Table Rows
        for obj in queryset:
            row = []
            for i, field in enumerate(fields):
                value = getattr(obj, field[0], "")

                if isinstance(value, bool):
                    value = "نعم" if value else "لا"
                elif isinstance(value, (int, float, Decimal)):
                    value = "{:,.2f}".format(value)
                elif isinstance(value, (datetime, date)):
                    value = value.strftime("%Y-%m-%d")
                elif hasattr(obj, f"get_{field[0]}_display"):
                    value = getattr(obj, f"get_{field[0]}_display")()  # Choice fields

                value = str(value) if value else ""

                # Fix Arabic text order
                if any("\u0600" <= c <= "\u06FF" for c in value):
                    value = get_display(arabic_reshaper.reshape(value))

                row.append(value)
                max_col_lengths[i] = max(max_col_lengths[i], len(value))  # Track longest text

            table_data.append(row)

        # Calculate column widths dynamically
        total_width = 780  # Approximate A4 width in landscape mode (in points)
        min_width = 80  # Minimum column width
        max_width = 220  # Maximum column width
        scale_factor = total_width / sum(max_col_lengths)  # Normalize sizes

        col_widths = [max(min_width, min(int(l * scale_factor), max_width)) for l in max_col_lengths]

        # Create Table
        table = Table(table_data, colWidths=col_widths)


        # Create Table
        # table = Table(table_data, colWidths=[len(fields)-20] * len(fields))
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 0), (-1, -1), "Janna"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ]))

        # **Title (Centered and Bold)**
        title_text = get_display(arabic_reshaper.reshape(f"تقرير {queryset.model._meta.verbose_name_plural}"))
        title_style = ParagraphStyle(name="Title", fontName="Janna", fontSize=16, alignment=1, spaceAfter=20)

        # **Subtitle (Smaller text under the title)**
        subtitle_text = get_display(arabic_reshaper.reshape(f"قائمة {queryset.model._meta.verbose_name_plural} المصدرة من النظام"))
        subtitle_style = ParagraphStyle(name="Subtitle", fontName="Janna", fontSize=10, alignment=1, textColor=colors.grey)

        elements.append(Paragraph(title_text, title_style))
        elements.append(Spacer(1, 20))
        elements.append(table)
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(subtitle_text, subtitle_style))


        doc.build(elements)

        # Get PDF content from buffer
        pdf_content = buffer.getvalue()
        buffer.close()

        # Use `prepare_export_response` for consistent file handling
        return self.prepare_export_response(pdf_content, "pdf", model_name)

    def export_by_fpdf(self, queryset):
        """Generate a properly formatted PDF with RTL Arabic support."""
        if not queryset.exists():
            return HttpResponse("No data to export.", content_type="text/plain")

        pdf = FPDF(orientation="L", unit="mm", format="A4")
        pdf.add_page()

        # Set Arabic font
        font_path = os.path.abspath(os.path.join("static", "fonts", "Janna LT Bold/Janna LT Bold.ttf"))
        if not os.path.exists(font_path):
            return HttpResponse("Font file not found.", content_type="text/plain")
        pdf.add_font("ArabicFont", "", font_path, uni=True)

        # Table Headers (Right-aligned for Arabic)
        pdf.set_font("ArabicFont", size=14)
        pdf.set_text_color(33, 37, 41)
        pdf.set_fill_color(222, 226, 230)
        pdf.set_draw_color(177, 179, 180)
        title = get_display(arabic_reshaper.reshape(f"قائمة {queryset.model._meta.verbose_name_plural}"))
        pdf.cell(200, 10, title, ln=True, align='R')
        pdf.set_font("ArabicFont", size=10)
        # Get model field names dynamically
        fields = [[field.name, field.verbose_name] for field in queryset.model._meta.fields[-1:1:-1]]
        column_width = (pdf.w - 20) / len(fields)
        # Table Headers
        for field in fields:
            pdf.cell(column_width, 10, get_display(arabic_reshaper.reshape(field[1])), border=0, fill=1, align="C")
        pdf.ln()

        pdf.set_font("ArabicFont", size=8)
        # Table Rows
        y = y2 = pdf.y
        for i in range(20):
            for obj in queryset:
                for i, field in enumerate(fields[:3]):
                    value = getattr(obj, field[0], "")

                    if isinstance(value, bool):
                        value = "نعم" if value else "لا"
                    elif isinstance(value, (int, float, Decimal)):
                        value = "{:,.2f}".format(value)
                    elif isinstance(value, (datetime, date)):
                        value = value.strftime("%Y-%m-%d")
                    elif hasattr(obj, f"get_{field[0]}_display"):
                        value = getattr(obj, f"get_{field[0]}_display")() # Choice fields

                    value = str(value) if value else ""

                    # Fix Arabic text order
                    if any("\u0600" <= c <= "\u06FF" for c in value):
                        reshaped_text = arabic_reshaper.reshape(value)
                        value = get_display(reshaped_text)

                    # pdf.cell(column_width, 10, value, border=0, align="R")
                    pdf.set_xy(column_width * i + 10, y)  # Reset position for wrapping
                    y0 = pdf.get_y()
                    pdf.multi_cell(column_width, 10, value, border=0, align="R")
                    y1 = pdf.get_y()
                    if y0<y1:
                        y1 = y1
                    else:
                        y1 = y1 + pdf.h
                        pdf.set_y(y- pdf.h)
                    y2 = max(y2, y1)
                    # print(y,'-',y0,y1,'-',y2,'-',pdf.get_y(),)
                y = y2 = y2 % pdf.h
                # print("---------------")
                # print(y, pdf.h, y2)
                # print("---------------")
                pdf.line(10, y, pdf.w - 10, y)
                pdf.ln()

        content = pdf.output(dest="S").encode("latin1", "replace")
        file_type = "pdf"
        model_name = queryset.model._meta.model_name
        return self.prepare_export_response(content, file_type, model_name)


# Loan Views
class LoanListView(PermissionRequiredMixin, ExportMixin, ListView):
    model = Loan
    template_name = 'loans/loan_list.html'
    context_object_name = 'loans'
    permission_required = 'core.view_loan'

    def get_queryset(self):
        queryset = super().get_queryset()
        owner_filter = self.request.GET.get('owner')
        loan_type_filter = self.request.GET.get('loan_type')
        date_filter = self.request.GET.get('date')

        if owner_filter:
            queryset = queryset.filter(loan_owner=owner_filter)
        if loan_type_filter:
            queryset = queryset.filter(loan_type=loan_type_filter)
        if date_filter:
            queryset = queryset.filter(date=date_filter)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = LoanFilterForm(self.request.GET)
        return context

class LoanDetailView(PermissionMixin, DetailView):
    model = Loan
    template_name = 'loans/loan_detail.html'
    permission_required = 'core.view_loan'

class LoanCreateView(PermissionMixin, CreateView):
    model = Loan
    form_class = LoanForm
    # fields = ['user', 'loan_type', 'amount', 'paid_amount', 'date']
    template_name = 'loans/loan_form.html'
    success_url = reverse_lazy('loan_list')
    permission_required = 'core.add_loan'

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'تم إضافة السلفة بنجاح.')
        return response

    def form_invalid(self, form):
        messages.error(self.request, 'حدث خطأ أثناء إضافة السلفة. يرجى المحاولة مرة أخرى.')
        return super().form_invalid(form)

    def get_form(self):
        form = super().get_form()
        form.fields['remaining_amount'].widget = HiddenInput()
        return form

class LoanUpdateView(PermissionMixin, UpdateView):
    model = Loan
    form_class = LoanForm
    template_name = 'loans/loan_form.html'
    success_url = reverse_lazy('loan_list')
    permission_required = 'core.change_loan'

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'تم تحديث السلفة بنجاح.')
        return response

    def form_invalid(self, form):
        messages.error(self.request, 'حدث خطأ أثناء تحديث السلفة. يرجى المحاولة مرة أخرى.')
        return super().form_invalid(form)

    def get_form(self):
        form = super().get_form()
        # لجعل حقل صاحب السلف غير قابل للتعديل عند التحديث
        # form.fields['loan_owner'].disabled = True
        form.fields['remaining_amount'].disabled = True
        return form

class LoanDeleteView(PermissionMixin, DeleteView):
    model = Loan
    template_name = 'loans/loan_confirm_delete.html'
    success_url = reverse_lazy('loan_list')
    permission_required = 'core.delete_loan'

    def delete(self, request, *args, **kwargs):
        response = super().delete(request, *args, **kwargs)
        messages.success(request, 'تم حذف السلفة بنجاح.')
        return response


# Covenant Views
class CovenantListView(PermissionMixin, ExportMixin, ListView):
    model = Covenant
    form = CovenantForm()
    template_name = 'covenants/covenant_list.html'
    context_object_name = 'covenants'
    permission_required = 'core.view_covenant'

    def get_queryset(self):
        queryset = super().get_queryset()
        owner_filter = self.request.GET.get('owner')
        covenant_type_filter = self.request.GET.get('covenant_type')
        date_filter = self.request.GET.get('date')

        if owner_filter:
            queryset = queryset.filter(covenant_owner=owner_filter)
        if covenant_type_filter:
            queryset = queryset.filter(covenant_type=covenant_type_filter)
        if date_filter:
            queryset = queryset.filter(date=date_filter)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = CovenantFilterForm(self.request.GET)
        return context

class CovenantDetailView(PermissionMixin, DetailView):
    model = Covenant
    template_name = 'covenants/covenant_detail.html'
    permission_required = 'core.view_covenant'

class CovenantCreateView(PermissionMixin, CreateView):
    model = Covenant
    form_class = CovenantForm
    # fields = ['user', 'covenant_type', 'amount', 'paid_amount', 'date']
    template_name = 'covenants/covenant_form.html'
    success_url = reverse_lazy('covenant_list')
    permission_required = 'core.add_covenant'

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'تم إضافة العهد بنجاح.')
        return response

    def form_invalid(self, form):
        messages.error(self.request, 'حدث خطأ أثناء إضافة العهد. يرجى المحاولة مرة أخرى.')
        return super().form_invalid(form)

    def get_form(self):
        form = super().get_form()
        form.fields['remaining_amount'].widget = HiddenInput()
        return form

class CovenantUpdateView(PermissionMixin, UpdateView):
    model = Covenant
    form_class = CovenantForm
    template_name = 'covenants/covenant_form.html'
    success_url = reverse_lazy('covenant_list')
    permission_required = 'core.change_covenant'

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'تم تحديث العهد بنجاح.')
        return response

    def form_invalid(self, form):
        messages.error(self.request, 'حدث خطأ أثناء تحديث العهد. يرجى المحاولة مرة أخرى.')
        return super().form_invalid(form)

    def get_form(self):
        form = super().get_form()
        # لجعل حقل صاحب السلف غير قابل للتعديل عند التحديث
        # form.fields['covenant_owner'].disabled = True
        form.fields['remaining_amount'].disabled = True
        return form

class CovenantDeleteView(PermissionMixin, DeleteView):
    model = Covenant
    template_name = 'covenants/covenant_confirm_delete.html'
    success_url = reverse_lazy('covenant_list')
    permission_required = 'core.delete_covenant'

    def delete(self, request, *args, **kwargs):
        response = super().delete(request, *args, **kwargs)
        messages.success(request, 'تم حذف العهد بنجاح.')
        return response


# Client Views
class ClientListView(PermissionMixin, ExportMixin, ListView):
    model = Client
    template_name = 'clients/client_list.html'
    context_object_name = 'clients'
    permission_required = 'core.view_client'

    def get_queryset(self):
        queryset = super().get_queryset()
        name_filter = self.request.GET.get('name')
        type_filter = self.request.GET.get('type')
        category_filter = self.request.GET.get('category')
        date_filter = self.request.GET.get('date')

        if name_filter:
            queryset = queryset.filter(name=name_filter)
        if type_filter:
            queryset = queryset.filter(type=type_filter)
        if category_filter:
            queryset = queryset.filter(category=category_filter)
        if date_filter:
            queryset = queryset.filter(date=date_filter)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = ClientFilterForm(self.request.GET)
        return context

class ClientDetailView(PermissionMixin, DetailView):
    model = Client
    template_name = 'clients/client_detail.html'
    permission_required = 'core.view_client'

    def get_object(self):
        client = super().get_object()
        # Debugging line
        filtered_revenue = client.budget_revenues.filter(remaining_amount__gt=0)
        # Add the filtered revenues to the client object
        client.filtered_revenue = filtered_revenue
        return client

class ClientDeleteView(PermissionMixin, DeleteView):
    model = Client
    template_name = 'clients/client_confirm_delete.html'
    success_url = reverse_lazy('client_list')
    permission_required = 'core.delete_client'

    def delete(self, request, *args, **kwargs):
        response = super().delete(request, *args, **kwargs)
        messages.success(request, 'تم حذف العميل بنجاح.')
        return response

class ClientFormView(PermissionMixin, CreateView, UpdateView):
    model = Client
    form_class = ClientForm
    template_name = 'clients/client_form.html'
    success_url = reverse_lazy('client_list')

    def get_object(self, queryset=None):
        """Return None for create, existing instance for update."""
        pk = self.kwargs.get('pk')
        if pk:
            return get_object_or_404(Client, pk=pk)
        return None

    def has_permission(self):
        """Check user permissions dynamically based on action."""
        client = self.get_object()
        if client:
            return self.request.user.has_perm('core.change_client')
        return self.request.user.has_perm('core.add_client')

    def get_context_data(self, **kwargs):
        """Add phone, email, and document formsets to the context."""
        context = super().get_context_data(**kwargs)
        instance = self.get_object()

        # Initialize formsets
        if self.request.POST:
            context['phone_formset'] = ClientPhoneFormSet(self.request.POST, instance=instance)
            context['email_formset'] = ClientEmailFormSet(self.request.POST, instance=instance)
            context['document_formset'] = ClientDocumentFormSet(self.request.POST, self.request.FILES, instance=instance)

        else:
            context['phone_formset'] = ClientPhoneFormSet(instance=instance)
            context['email_formset'] = ClientEmailFormSet(instance=instance)
            context['document_formset'] = ClientDocumentFormSet(instance=instance)

        return context

    def form_valid(self, form):
        """Validate main form and related formsets."""
        context = self.get_context_data()
        phone_formset = context['phone_formset']
        email_formset = context['email_formset']
        document_formset = context['document_formset']

        if form.is_valid() and phone_formset.is_valid() and email_formset.is_valid() and document_formset.is_valid():
            self.object = form.save()

            phone_formset.instance = self.object
            email_formset.instance = self.object
            document_formset.instance = self.object

            phone_instances = phone_formset.save(commit=False)
            email_instances = email_formset.save(commit=False)
            document_instances = document_formset.save(commit=False)

            for form in phone_formset:
                if form.cleaned_data.get('DELETE', False):  # Check delete flag
                    form.instance.delete()
                else:
                    form.save()

            for form in email_formset:
                if form.cleaned_data.get('DELETE', False):  # Check delete flag
                    form.instance.delete()
                else:
                    form.save()

            for form in document_formset:
                if form.cleaned_data.get('DELETE', False):  # Check delete flag
                    form.instance.delete()
                else:
                    form.save()

            phone_formset.save_m2m()
            email_formset.save_m2m()
            document_formset.save_m2m()

            messages.success(self.request, "تم حفظ البيانات بنجاح.")
            return super().form_valid(form)  # Ensure this is always returned
        else:
            # Show specific validation errors
            if not phone_formset.is_valid():
                messages.error(self.request, f"حدث خطأ أثناء حفظ أرقام الهاتف: {phone_formset.errors}")
            if not email_formset.is_valid():
                messages.error(self.request, f"حدث خطأ أثناء حفظ البريد الإلكتروني: {email_formset.errors}")
            if not document_formset.is_valid():
                messages.error(self.request, f"حدث خطأ أثناء حفظ المستندات: {document_formset.errors}")

            return self.form_invalid(form)  # Ensure this is also consistently returned

    def form_invalid(self, form):
        """Handle invalid form submissions."""
        messages.error(self.request, "حدث خطأ أثناء حفظ البيانات. يرجى التحقق من المدخلات.")
        return super().form_invalid(form)


# Budget Views
class BudgetView(PermissionMixin, ExportMixin, View):
    model = Budget
    template_name = 'budgets/budget_list.html'
    permission_required = 'core.view_budget'

    def get(self, request, *args, **kwargs):
        budgets = Budget.objects.filter(users=request.user) if not request.user.is_superuser else Budget.objects.all()
        form = BudgetForm()
        budget = None
    
        # Apply additional filters from GET parameters
        name_filter = request.GET.get('name')
        status_filter = request.GET.get('status')
        date_filter = request.GET.get('date')

        if name_filter:
            budgets = budgets.filter(name=name_filter)
        if status_filter:
            budgets = budgets.filter(status=status_filter)
        if date_filter:
            budgets = budgets.filter(date__icontains=date_filter)

        if 'pk' in kwargs:
            budget = get_object_or_404(Budget, pk=kwargs['pk'])
            form = BudgetForm(instance=budget)
    
        if "export" in request.GET:
            export_format = request.GET.get('format', 'excel')  
            if export_format == 'pdf':
                return self.export_to_pdf(budgets)
            return self.export_to_excel(budgets)
    
        filter_form = BudgetFilterForm(request.GET)
        return render(request, self.template_name, {
            'budgets': budgets,
            'form': form,
            'budget': budget,
            'filter_form': filter_form
        })
    
    def post(self, request, *args, **kwargs):
        if 'pk' in kwargs:
            budget = get_object_or_404(Budget, pk=kwargs['pk'])
            form = BudgetForm(request.POST, instance=budget)
        else:
            form = BudgetForm(request.POST)

        if form.is_valid():
            budget = form.save(commit=False)  # Do not save immediately            
            
            # Ensure 'date' is properly updated
            if 'date' in form.cleaned_data:
                budget.date = form.cleaned_data['date']
            
            budget.save()
            
            # Handle ManyToManyField update
            if not request.user.is_superuser:
                budget.users.add(request.user) 

            form.save_m2m()
            messages.success(request, "تم حفظ الميزانية بنجاح!")
            return redirect('budget_list')
    
        messages.error(request, f"حدث خطأ أثناء إرسال النموذج. يرجى المحاولة مرة أخرى.")
        
        budgets = Budget.objects.filter(users=request.user) if not request.user.is_superuser else Budget.objects.all()
        return render(request, self.template_name, {
            'budgets': budgets,
            'form': form
        })

class BudgetDetailView(PermissionMixin, ExportMixin, DetailView):
    model = Budget
    template_name = "budgets/budget_detail.html"
    context_object_name = "budget"
    permission_required = "core.view_budget"

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        context = self.get_context_data()

        # Handle export request
        if "export" in self.request.GET:
            export_format = self.request.GET.get("format", "excel")
            if export_format == "pdf":
                return self.export_to_pdf(context)
            else:
                return self.export_to_excel(context)
        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Retrieve expenses and revenues
        expenses = BudgetExpense.objects.filter(budget=self.object)
        revenues = BudgetRevenue.objects.filter(budget=self.object)

        # Initialize filter forms
        expense_filter_form = BudgetExpenseFilterForm(self.request.GET)
        revenue_filter_form = BudgetRevenueFilterForm(self.request.GET)

        # Apply expense filters
        if expense_filter_form.is_valid():
            filters = {k.replace('expense_', ''): v for k, v in expense_filter_form.cleaned_data.items() if v}
            expenses = expenses.filter(**filters)

        # Apply revenue filters
        if revenue_filter_form.is_valid():
            filters = {k.replace('revenue_', ''): v for k, v in revenue_filter_form.cleaned_data.items() if v}
            revenues = revenues.filter(**filters)
        
        # Aggregate totals
        expense_totals = expenses.aggregate(
            total_expenses_estimated=Sum("estimated_amount", default=0),
            total_expenses_actual=Sum("actual_amount", default=0),
            total_expenses_paid=Sum("paid_amount", default=0),
            total_expenses_remaining=Sum("remaining_amount", default=0),
        )
        revenue_totals = revenues.aggregate(
            total_revenues_estimated=Sum("estimated_amount", default=0),
            total_revenues_actual=Sum("actual_amount", default=0),
            total_revenues_paid=Sum("paid_amount", default=0),
            total_revenues_remaining=Sum("remaining_amount", default=0),
        )

        # Calculate surplus
        estimated_surplus = revenue_totals["total_revenues_estimated"] - expense_totals["total_expenses_estimated"]
        actual_surplus = revenue_totals["total_revenues_actual"] - expense_totals["total_expenses_actual"]

        # Format numerical values
        format_currency = lambda x: "{:,.2f}".format(x)
        formatted_data = {key: format_currency(value) for key, value in {
            **expense_totals, **revenue_totals,
            "estimated_surplus": estimated_surplus,
            "actual_surplus": actual_surplus,
        }.items()}

        # Format values for display
        context.update({
            'expenses': expenses,
            'expenses': expenses if self.request.user.has_perm('core.view_budgetexpense') else None,
            'revenues': revenues if self.request.user.has_perm('core.view_budgetrevenue') else None,
            'expense_filter_form': expense_filter_form,
            'revenue_filter_form': revenue_filter_form,
            **formatted_data,
        })
        return context

    def export_to_pdf(self, context):
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=10, leftMargin=10, topMargin=20, bottomMargin=20)
    
        # Register Arabic font
        font_path = os.path.abspath(os.path.join(settings.STATIC_ROOT, "fonts", "Janna LT Bold", "Janna LT Bold.ttf"))
        pdfmetrics.registerFont(TTFont("Janna", font_path))
    
        # Styles
        title_style = ParagraphStyle(name="Title", fontName="Janna", fontSize=16, alignment=1, spaceAfter=20)
        subtitle_style = ParagraphStyle(name="Subtitle", fontName="Janna", fontSize=12, alignment=2, textColor=colors.grey)
        text_style = ParagraphStyle(name="Text", fontName="Janna", fontSize=10, alignment=2, spaceAfter=5)  # Right-aligned for Arabic text
        table_style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 0), (-1, -1), "Janna"),
            ("FONTSIZE", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ])
    
        budget_details = [
            [
             Paragraph(get_display(arabic_reshaper.reshape(context["budget"].status)), text_style),
             Paragraph(get_display(arabic_reshaper.reshape("الحالة:")), subtitle_style), 
             Paragraph(get_display(arabic_reshaper.reshape("الوصف:")), subtitle_style)
            ], [
             Paragraph(context["budget"].date.strftime("%m-%Y"), text_style), 
             Paragraph(get_display(arabic_reshaper.reshape("تاريخ الموازنة:")), subtitle_style),
             Paragraph(get_display(arabic_reshaper.reshape(context["budget"].description)), text_style),
            ]
        ]
        
        budget_table = Table(budget_details, colWidths=[150, 90, 450])
        budget_name = context["budget"].name
        elements = [
            Paragraph(get_display(arabic_reshaper.reshape(f"تفاصيل الموازانة: {budget_name}")), title_style),
            Spacer(1, 20),
            budget_table,
            Spacer(1, 20),
            Paragraph(get_display(arabic_reshaper.reshape("قائمة المصاريف والإيرادات")), ParagraphStyle(name="Subtitle", fontName="Janna", fontSize=12, alignment=1, textColor=colors.grey)),
            Spacer(1, 10),
        ]

        revenues = context.get("revenues") or []
        expenses = context.get("expenses") or []
            
        revenues_fields = [[field.name, field.verbose_name] for field in BudgetRevenue._meta.get_fields()[-1:1:-1]] if revenues else []
        expenses_fields = [[field.name, field.verbose_name] for field in BudgetExpense._meta.get_fields()[-2:1:-1]] if expenses else []
        fields = revenues_fields + expenses_fields 
        headers = [get_display(arabic_reshaper.reshape(field[1])) for field in fields]
        table_data = [headers]  # Table header

        max_col_lengths = [len(header) for header in headers]
        # Table Rows
        def add_data(obj, fields, row):
            for i, field in enumerate(fields):
                value = getattr(obj, field[0], "")

                if isinstance(value, bool):
                    value = "نعم" if value else "لا"
                elif isinstance(value, (int, float, Decimal)):
                    value = "{:,.2f}".format(value)
                elif isinstance(value, (datetime, date)):
                    value = value.strftime("%Y-%m-%d")
                elif hasattr(obj, f"get_{field[0]}_display"):
                    value = getattr(obj, f"get_{field[0]}_display")()  # Choice fields

                value = str(value) if value else ""

                # Fix Arabic text order
                if any("\u0600" <= c <= "\u06FF" for c in value):
                    value = get_display(arabic_reshaper.reshape(value))

                row.append(value)
                max_col_lengths[i] = max(max_col_lengths[i], len(value))  # Track longest text

        if len(headers):
            for i in range(max(len(list(revenues)), len(list(expenses)))):
                row = []
                if i < len(list(revenues)):
                    obj = revenues[i]
                    add_data(obj, revenues_fields, row)
                if i < len(list(expenses)):
                    obj = expenses[i]
                    add_data(obj, expenses_fields, row)

                table_data.append(row)

            # Calculate column widths dynamically
            total_width = 780  # Approximate A4 width in landscape mode (in points)
            min_width = 20  # Minimum column width
            max_width = 220  # Maximum column width
            scale_factor = total_width / max(1, sum(max_col_lengths)) # Normalize sizes

            col_widths = [max(min_width, min(int(l * scale_factor), max_width)) for l in max_col_lengths]

            # Create Table
            table = Table(table_data, colWidths=col_widths)
            table.setStyle(table_style)
            elements.append(table)
            elements.append(Spacer(1, 10))

        # Build PDF
        doc.build(elements)
        pdf_content = buffer.getvalue()
        buffer.close()
    
        return self.prepare_export_response(pdf_content, "pdf", self.model._meta.model_name)

    def export_to_excel(self, context):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Budget Details"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
    
        # Budget details
        budget = context["budget"]
        for field in budget._meta.get_fields()[4:-1]:
            sheet.append([str(getattr(budget, field.name, "")), field.verbose_name])

        sheet.append([])  # Empty row for spacing
    
        # Create **Revenue & Expense Table Headers**
        revenue_headers = [field.verbose_name if hasattr(field, 'verbose_name') else field.name for field in BudgetRevenue._meta.get_fields()[-1:1:-1]]
        expense_headers = [field.verbose_name if hasattr(field, 'verbose_name') else field.name for field in BudgetExpense._meta.get_fields()[-1:1:-1]]
    
        header_row = revenue_headers + [""] + expense_headers
        sheet.append(header_row)
    
        # Apply styles to headers
        for col_idx, col in enumerate(header_row, start=1):
            cell = sheet.cell(row=sheet.max_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
        # Fetching transactions
        expenses = context.get("expenses", [])
        revenues = context.get("revenues", [])
    
        max_len = max(len(revenues), len(expenses))
    
        for i in range(max_len):
            row = []
    
            # Revenue Columns
            if i < len(revenues):
                row.extend([str(getattr(revenues[i], field.name, "")) for field in BudgetRevenue._meta.get_fields()[-1:1:-1]])
            else:
                row.extend([""] * len(revenue_headers))
    
            row.append("")  # Spacer column
    
            # Expense Columns
            if i < len(expenses):
                row.extend([str(getattr(expenses[i], field.name, "")) for field in BudgetExpense._meta.get_fields()[-1:1:-1]])
            else:
                row.extend([""] * len(expense_headers))
    
            sheet.append(row)
    
        # Adjust column widths dynamically
        for col in sheet.columns:
            max_length = 0
            col_letter = col[0].column_letter  # Get the column letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[col_letter].width = adjusted_width
    
        # Create response
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        content = output.read()
        return self.prepare_export_response(content, "xlsx", self.model._meta.model_name)

class BudgetRevenueCreateView(PermissionMixin, CreateView):
    model = BudgetRevenue
    form_class = BudgetRevenueForm
    # fields = ['name', 'email', 'phone', 'type', 'category', 'join_date', 'balance']
    template_name = 'budgets/budget_revenue_form.html'
    success_url = reverse_lazy('budget_list')
    permission_required = 'core.add_budgetrevenue'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['budget'] = get_object_or_404(Budget, pk=self.kwargs['budget_pk'])
        return context

    def get_success_url(self):
        return reverse_lazy('budget_detail', kwargs={'pk': self.kwargs['budget_pk']})

    def form_valid(self, form):
        budget = get_object_or_404(Budget, pk=self.kwargs['budget_pk'])
        form.instance.budget = budget
        response = super().form_valid(form)
        messages.success(self.request, 'تم إضافة الايراد بنجاح.')
        return response

    def form_invalid(self, form):
        messages.error(self.request, 'حدث خطأ أثناء إضافة الايراد. يرجى المحاولة مرة أخرى.')
        return super().form_invalid(form)

class BudgetRevenueUpdateView(PermissionMixin, UpdateView):
    model = BudgetRevenue
    form_class = BudgetRevenueForm
    template_name = 'budgets/budget_revenue_form.html'
    success_url = reverse_lazy('budget_list')
    permission_required = 'core.change_budgetrevenue'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['budget'] = get_object_or_404(Budget, pk=self.kwargs['budget_pk'])
        return context

    def get_success_url(self):
        return reverse_lazy('budget_detail', kwargs={'pk': self.kwargs['budget_pk']})

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'تم تحديث الايراد بنجاح.')
        return response

    def form_invalid(self, form):
        messages.error(self.request, 'حدث خطأ أثناء تحديث الايراد. يرجى المحاولة مرة أخرى.')
        return super().form_invalid(form)

class BudgetRevenuesChangeView(PermissionMixin, View):
    template_name = 'budgets/budget_revenues_change.html'
    permission_required = 'core.change_budgetrevenue'

    def get(self, request, budget_pk):
        # Get the budget instance
        budget = get_object_or_404(Budget, pk=budget_pk)

        # Get the existing revenues for the budget
        revenues = BudgetRevenue.objects.filter(budget=budget)

        # Create a form instance for each revenue item
        revenue_forms = [
            BudgetRevenueForm(instance=revenue, prefix=f"revenue_{revenue.pk}")
            for revenue in revenues
        ]


        # Add a form to add a new revenue
        revenue_form = BudgetRevenueForm()

        context = {
            'budget': budget,
            'revenue_forms': revenue_forms,
            'revenue_form': revenue_form,
        }

        return render(request, self.template_name, context)

    def post(self, request, budget_pk):
        # Get the budget instance
        budget = get_object_or_404(Budget, pk=budget_pk)

        # Check if a delete action is triggered
        if "delete_revenue" in request.POST:
            revenue_id = request.POST.get("delete_revenue")
            revenue = get_object_or_404(BudgetRevenue, pk=revenue_id, budget=budget)
            revenue.delete()
            return redirect('budget_revenues_change', budget_pk=budget.pk)

        # Get the existing revenues for the budget
        revenues = BudgetRevenue.objects.filter(budget=budget)

        # Create form instances for each revenue item
        revenue_forms = [
            BudgetRevenueForm(
                request.POST,
                instance=revenue,
                prefix=f"revenue_{revenue.pk}"
            )
            for revenue in revenues
        ]

        # Create new revenue form to add new items
        revenue_form = BudgetRevenueForm(request.POST)

        # Update existing revenues
        for form in revenue_forms:
            if form.is_valid():
                form.save()

        # Add new revenue if valid
        if revenue_form.is_valid():
            new_revenue = revenue_form.save(commit=False)
            new_revenue.budget = budget
            new_revenue.save()

        # Redirect back to the same page with updated revenues
        return redirect('budget_revenues_change', budget_pk=budget.pk)

class BudgetExpensesChangeView(PermissionMixin, View):
    template_name = 'budgets/budget_expenses_change.html'
    permission_required = 'core.change_budgetexpense'

    def get(self, request, budget_pk):
        # Get the budget instance
        budget = get_object_or_404(Budget, pk=budget_pk)

        # Get the existing expenses for the budget
        expenses = BudgetExpense.objects.filter(budget=budget)

        # Create a form instance for each expense item with unique prefixes
        expense_forms = [
            BudgetExpenseForm(instance=expense, prefix=f"expense_{expense.pk}")
            for expense in expenses
        ]

        # Add a form to add a new expense
        expense_form = BudgetExpenseForm(prefix="new_expense")

        context = {
            'budget': budget,
            'expense_forms': expense_forms,
            'expense_form': expense_form,
        }

        return render(request, self.template_name, context)

    def post(self, request, budget_pk):
        # Get the budget instance
        budget = get_object_or_404(Budget, pk=budget_pk)

        # Check if a delete action is triggered
        if "delete_expense" in request.POST:
            expense_id = request.POST.get("delete_expense")
            expense = get_object_or_404(BudgetExpense, pk=expense_id, budget=budget)
            expense.delete()
            return redirect('budget_expenses_change', budget_pk=budget.pk)

        # Get the existing expenses for the budget
        expenses = BudgetExpense.objects.filter(budget=budget)

        # Create form instances for each expense item with unique prefixes
        expense_forms = [
            BudgetExpenseForm(
                request.POST,
                instance=expense,
                prefix=f"expense_{expense.pk}"
            )
            for expense in expenses
        ]

        # Create a new expense form to add new items
        expense_form = BudgetExpenseForm(request.POST, prefix="new_expense")

        # Flag to check if all forms are valid
        all_valid = all(form.is_valid() for form in expense_forms)

        # Update existing expenses if all forms are valid
        if all_valid:
            for form in expense_forms:
                form.save()

        # Add a new expense if valid
        if expense_form.is_valid():
            new_expense = expense_form.save(commit=False)
            new_expense.budget = budget
            new_expense.save()

        # Redirect back to the same page with updated expenses
        return redirect('budget_expenses_change', budget_pk=budget.pk)

class BudgetExpenseCreateView(PermissionMixin, CreateView):
    model = BudgetExpense
    form_class = BudgetExpenseForm
    template_name = 'budgets/budget_expense_form.html'
    success_url = reverse_lazy('budget_list')
    permission_required = 'core.add_budgetexpense'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['budget'] = get_object_or_404(Budget, pk=self.kwargs['budget_pk'])
        return context

    def get_success_url(self):
        return reverse_lazy('budget_detail', kwargs={'pk': self.kwargs['budget_pk']})

    def form_valid(self, form):
        budget = get_object_or_404(Budget, pk=self.kwargs['budget_pk'])
        form.instance.budget = budget
        response = super().form_valid(form)
        messages.success(self.request, 'تم إضافة المصروف بنجاح.')
        return response

    def form_invalid(self, form):
        messages.error(self.request, 'حدث خطأ أثناء إضافة المصروف. يرجى المحاولة مرة أخرى.')
        return super().form_invalid(form)

class BudgetExpenseUpdateView(PermissionMixin, UpdateView):
    model = BudgetExpense
    form_class = BudgetExpenseForm
    template_name = 'budgets/budget_expense_form.html'
    success_url = reverse_lazy('budget_list')
    permission_required = 'core.change_budgetexpense'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['budget'] = get_object_or_404(Budget, pk=self.kwargs['budget_pk'])
        return context

    def get_success_url(self):
        return reverse_lazy('budget_detail', kwargs={'pk': self.kwargs['budget_pk']})

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'تم تحديث المصروف بنجاح.')
        return response

    def form_invalid(self, form):
        messages.error(self.request, 'حدث خطأ أثناء تحديث المصروف. يرجى المحاولة مرة أخرى.')
        return super().form_invalid(form)


# Fund Views
class FundView(PermissionMixin, ExportMixin, View):
    template_name = 'funds/fund_list.html'
    permission_required = 'core.view_fund'

    def get(self, request, *args, **kwargs):
        # عرض جميع الصناديق إذا كان المستخدم مشرفًا، وإلا يعرض الصناديق الخاصة به فقط
        if request.user.has_perm('core.viewprivate_funds') and request.user.has_perm('core.viewall_funds'):
            funds = Fund.objects.all()
        elif request.user.has_perm('core.viewprivate_funds'):
            funds = Fund.objects.filter(user=request.user)
        elif request.user.has_perm('core.viewall_funds'):
            funds = Fund.objects.filter(is_private=False)
        else:
            funds = Fund.objects.filter(user=request.user, is_private=False)

        # Apply additional filters from GET parameters
        user_filter = request.GET.get('user')
        name_filter = request.GET.get('name')
        date_filter = request.GET.get('date')

        if user_filter:
            funds = funds.filter(user=user_filter)
        if name_filter:
            funds = funds.filter(name=name_filter)
        if date_filter:
            funds = funds.filter(date__icontains=date_filter)

        # If an export is requested, send the filtered data
        if "export" in request.GET:
            export_format = request.GET.get('format', 'excel')  # Default to 'excel' if no format is provided
            if export_format == 'pdf':
                return self.export_to_pdf(funds)
            return self.export_to_excel(funds)

        fund = None
        if 'pk' in kwargs:
            fund = get_object_or_404(funds, pk=kwargs['pk'])  # التأكد من أن المستخدم يمكنه رؤية الصندوق
            form = FundForm(instance=fund)
        else:
            initial_data = {'user': request.user} if not request.user.is_superuser else {}
            form = FundForm(initial=initial_data)

        # Add filtering functionality to the context
        filter_form = FundFilterForm(request.GET)
        context = {
            'funds': funds,
            'form': form,
            'fund': fund,
            'filter_form': filter_form
        }
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        fund = None
        if 'pk' in kwargs:
            fund = get_object_or_404(Fund, pk=kwargs['pk'])
            form = FundForm(request.POST, instance=fund)
        else:
            form = FundForm(request.POST)

        if form.is_valid():
            fund = form.save(commit=False)
            if not request.user.is_superuser:
                fund.user = request.user  # تعيين المالك عند إنشاء الصندوق من قبل مستخدم عادي
            fund.save()
            return redirect('fund_list')  # Make sure 'fund_list' is the correct URL 
        
        messages.error(request, f"حدث خطأ أثناء إرسال النموذج. يرجى المحاولة مرة أخرى.")

        # إعادة تحميل الصناديق بعد الخطأ
        # عرض جميع الصناديق إذا كان المستخدم مشرفًا، وإلا يعرض الصناديق الخاصة به فقط
        if request.user.has_perm('core.viewprivate_funds') and request.user.has_perm('core.viewall_funds'):
            funds = Fund.objects.all()
        elif request.user.has_perm('core.viewprivate_funds'):
            funds = Fund.objects.filter(user=request.user)
        elif request.user.has_perm('core.viewall_funds'):
            funds = Fund.objects.filter(is_private=False)
        else:
            funds = Fund.objects.filter(user=request.user, is_private=False)

        # Reapply filters after form submission
        filter_form = FundFilterForm(request.GET)
        context = {
            'funds': funds,
            'form': form,
            'filter_form': filter_form
        }
        return render(request, self.template_name, context)

class FundDetailView(PermissionMixin, ExportMixin, DetailView):
    model = Fund
    template_name = 'funds/fund_detail.html'
    context_object_name = 'fund'
    permission_required = 'core.view_fund'

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        context = self.get_context_data()

        # Handle export request
        if "export" in self.request.GET:
            export_format = self.request.GET.get("format", "excel")
            if export_format == "pdf":
                return self.export_to_pdf(context)
            return self.export_to_excel(context)
        return self.render_to_response(context)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Retrieve expenses and revenues
        expenses = FundExpense.objects.filter(fund=self.object)
        revenues = FundRevenue.objects.filter(fund=self.object)

        # Initialize filter forms
        expense_filter_form = FundExpenseFilterForm(self.request.GET)
        revenue_filter_form = FundRevenueFilterForm(self.request.GET)

        # Apply expense filters
        if expense_filter_form.is_valid():
            filters = {k.replace('expense_', ''): v for k, v in expense_filter_form.cleaned_data.items() if v}
            expenses = expenses.filter(**filters)

        # Apply revenue filters
        if revenue_filter_form.is_valid():
            filters = {k.replace('revenue_', ''): v for k, v in revenue_filter_form.cleaned_data.items() if v}
            revenues = revenues.filter(**filters)
    
        # Aggregate totals for expenses and revenues
        total_expenses = expenses.aggregate(total=Sum('amount'))['total'] or 0
        total_revenues = revenues.aggregate(total=Sum('amount'))['total'] or 0

        # Format values for display
        context.update({
            'expenses': expenses if self.request.user.has_perm('core.view_budgetexpense') else None,
            'revenues': revenues if self.request.user.has_perm('core.view_budgetrevenue') else None,
            'total_expenses': "{:,.2f}".format(total_expenses),
            'total_revenues': "{:,.2f}".format(total_revenues),
            'expense_filter_form': expense_filter_form,
            'revenue_filter_form': revenue_filter_form,
        })
        return context

    def export_to_pdf(self, context):
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=10, leftMargin=10, topMargin=20, bottomMargin=20)
    
        # Register Arabic font
        font_path = os.path.abspath(os.path.join(settings.STATIC_ROOT, "fonts", "Janna LT Bold", "Janna LT Bold.ttf"))
        pdfmetrics.registerFont(TTFont("Janna", font_path))
    
        # Styles
        title_style = ParagraphStyle(name="Title", fontName="Janna", fontSize=16, alignment=1, spaceAfter=20)
        subtitle_style = ParagraphStyle(name="Subtitle", fontName="Janna", fontSize=12, alignment=2, textColor=colors.grey)
        text_style = ParagraphStyle(name="Text", fontName="Janna", fontSize=10, alignment=2, spaceAfter=5)  # Right-aligned for Arabic text
        table_style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 0), (-1, -1), "Janna"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ])
    
        fund_details = [
            [
             Paragraph(get_display(arabic_reshaper.reshape(context["fund"].formatted_current_balance())), text_style),
             Paragraph(get_display(arabic_reshaper.reshape("الرصيد الحالي:")), subtitle_style), 
             Paragraph(get_display(arabic_reshaper.reshape("الوصف:")), subtitle_style)
            ], [
             Paragraph(get_display(arabic_reshaper.reshape(context["fund"].formatted_opening_balance())), text_style),
             Paragraph(get_display(arabic_reshaper.reshape("الرصيد الافتتاحي:")), subtitle_style), 
             Paragraph(get_display(arabic_reshaper.reshape(context["fund"].description)), text_style),
            ], [
             Paragraph(context["fund"].date.strftime("%m-%Y"), text_style), 
             Paragraph(get_display(arabic_reshaper.reshape("تاريخ الصندوق:")), subtitle_style),
            ]
        ]
        
        fund_table = Table(fund_details, colWidths=[150, 110, 420])
        fund_name = context["fund"].name
        elements = [
            Paragraph(get_display(arabic_reshaper.reshape(f"تفاصيل الصندوق: {fund_name}")), title_style),
            Spacer(1, 20),
            fund_table,
            Spacer(1, 20),
            Paragraph(get_display(arabic_reshaper.reshape("قائمة المصاريف والإيرادات")), ParagraphStyle(name="Subtitle", fontName="Janna", fontSize=12, alignment=1, textColor=colors.grey)),
            Spacer(1, 10),
        ]

        revenues = context.get("revenues") or []
        expenses = context.get("expenses") or []
            
        revenues_fields = [[field.name, field.verbose_name] for field in FundRevenue._meta.get_fields()[-1:1:-1]] if revenues else []
        expenses_fields = [[field.name, field.verbose_name] for field in FundExpense._meta.get_fields()[-1:1:-1]] if expenses else []
        fields = revenues_fields + expenses_fields 
        headers = [get_display(arabic_reshaper.reshape(field[1])) for field in fields]
        table_data = [headers]  # Table header

        max_col_lengths = [len(header) for header in headers]
        # Table Rows
        def add_data(obj, fields, row):
            for i, field in enumerate(fields):
                value = getattr(obj, field[0], "")

                if isinstance(value, bool):
                    value = "نعم" if value else "لا"
                elif isinstance(value, (int, float, Decimal)):
                    value = "{:,.2f}".format(value)
                elif isinstance(value, (datetime, date)):
                    value = value.strftime("%Y-%m-%d")
                elif hasattr(obj, f"get_{field[0]}_display"):
                    value = getattr(obj, f"get_{field[0]}_display")()  # Choice fields

                value = str(value) if value else ""

                # Fix Arabic text order
                if any("\u0600" <= c <= "\u06FF" for c in value):
                    value = get_display(arabic_reshaper.reshape(value))

                row.append(value)
                max_col_lengths[i] = max(max_col_lengths[i], len(value))  # Track longest text

        if len(headers):
            for i in range(max(len(list(revenues)), len(list(expenses)))):
                row = []
                if i < len(list(revenues)):
                    obj = revenues[i]
                    add_data(obj, revenues_fields, row)
                if i < len(list(expenses)):
                    obj = expenses[i]
                    add_data(obj, expenses_fields, row)

                table_data.append(row)

            # Calculate column widths dynamically
            total_width = 780  # Approximate A4 width in landscape mode (in points)
            min_width = 20  # Minimum column width
            max_width = 220  # Maximum column width
            scale_factor = total_width / max(1, sum(max_col_lengths)) # Normalize sizes

            col_widths = [max(min_width, min(int(l * scale_factor), max_width)) for l in max_col_lengths]

            # Create Table
            table = Table(table_data, colWidths=col_widths)
            table.setStyle(table_style)
            elements.append(table)
            elements.append(Spacer(1, 10))

        # Build PDF
        doc.build(elements)
        pdf_content = buffer.getvalue()
        buffer.close()
    
        return self.prepare_export_response(pdf_content, "pdf", self.model._meta.model_name)

    def export_to_excel(self, context):
        # Create workbook and sheets
        wb = openpyxl.Workbook()
        ws_summary = wb.active
        ws_summary.title = "تفاصيل الصندوق"
        ws_transactions = wb.create_sheet(title="المعاملات المالية")
    
        # Set Arabic font
        font_path = os.path.abspath(os.path.join(settings.STATIC_ROOT, "fonts", "Janna LT Bold", "Janna LT Bold.ttf"))
        
        # Title style
        title_font = Font(name="Janna", size=14, bold=True)
        header_font = Font(name="Janna", size=12, bold=True)
        normal_font = Font(name="Janna", size=11)
    
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Alignment (Right for Arabic)
        right_align = Alignment(horizontal="right")
    
        # Fund Details
        fund = context["fund"]
        ws_summary.append(["تفاصيل الصندوق"])
        ws_summary["A1"].font = title_font
        ws_summary["A1"].alignment = right_align
    
        fund_data = [
            ["اسم الصندوق", get_display(arabic_reshaper.reshape(fund.name))],
            ["الرصيد الافتتاحي", fund.formatted_opening_balance()],
            ["الرصيد الحالي", fund.formatted_current_balance()],
            ["تاريخ الصندوق", fund.date.strftime("%Y-%m-%d")],
            ["الوصف", get_display(arabic_reshaper.reshape(fund.description or ""))],
        ]
    
        for row in fund_data:
            ws_summary.append(row)
    
        # Adjust column widths
        for col in ws_summary.columns:
            ws_summary.column_dimensions[col[0].column_letter].width = 30
    
        # Transactions Sheet (Revenues & Expenses)
        revenues = context.get("revenues", [])
        expenses = context.get("expenses", [])
    
        headers = []
        if revenues:
            headers = [[field.name, field.verbose_name] for field in FundRevenue._meta.get_fields()[-1:1:-1]]
        if expenses and not headers:
            headers = [[field.name, field.verbose_name] for field in FundExpense._meta.get_fields()[-1:1:-1]]
    
        if headers:
            ws_transactions.append([get_display(arabic_reshaper.reshape(field[1])) for field in headers])
    
            def add_transaction_data(obj, fields):
                row = []
                for field in fields:
                    value = getattr(obj, field[0], "")
    
                    if isinstance(value, bool):
                        value = "نعم" if value else "لا"
                    elif isinstance(value, (int, float, Decimal)):
                        value = "{:,.2f}".format(value)
                    elif isinstance(value, (datetime, date)):
                        value = value.strftime("%Y-%m-%d")
                    elif hasattr(obj, f"get_{field[0]}_display"):
                        value = getattr(obj, f"get_{field[0]}_display")()  # Choice fields
    
                    value = str(value) if value else ""
                    if any("\u0600" <= c <= "\u06FF" for c in value):
                        value = get_display(arabic_reshaper.reshape(value))
    
                    row.append(value)
                ws_transactions.append(row)
    
            for revenue in revenues:
                add_transaction_data(revenue, headers)
            for expense in expenses:
                add_transaction_data(expense, headers)
    
        # Apply styles to headers
        for col in ws_transactions[1]:
            col.font = header_font
            col.alignment = right_align
    
        # Adjust column widths
        for col in ws_transactions.columns:
            ws_transactions.column_dimensions[col[0].column_letter].width = 25
    
        # Save to BytesIO
        excel_stream = BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)
    
        return self.prepare_export_response(excel_stream.getvalue(), "xlsx", self.model._meta.model_name)
    
    def export_to_excel(self, context):
        # Create workbook and sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "التقرير المالي"

        # Styles
        title_font = Font(name="Janna", size=14, bold=True)
    
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Alignment (Right for Arabic)
        right_align = Alignment(horizontal="right")
    
        # 1️⃣ Add **Fund Details** at the Top
        fund = context["fund"]
        # ws.append(["تقرير مالي"])
        # ws["A1"].font = title_font
        # ws["A1"].alignment = right_align
    
        fund_data = [
            [fund.name, "اسم الصندوق"],
            [fund.formatted_opening_balance(), "الرصيد الافتتاحي"],
            [fund.formatted_current_balance(), "الرصيد الحالي"],
            [fund.date.strftime("%Y-%m-%d"), "تاريخ الصندوق"],
            [fund.description or "", "الوصف"],
        ]
    
        for row in fund_data:
            ws.append(row)
        
        ws.append([])  # Empty row for spacing
    
        # 2️⃣ Create **Revenue & Expense Table Headers**
        header_row = [
            "التاريخ", "المصدر", "المبلغ", "الوصف",   # Revenue Columns
            "",                          # Spacer
            "التاريخ", "المستلم", "المبلغ", "الوصف"  # Expense Columns
        ]
        ws.append([col if col else "" for col in header_row])
    
        # Apply styles to headers
        for col_idx, col in enumerate(header_row, start=1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = right_align
    
        # 3️⃣ Add **Revenue & Expense Data**
        revenues = context.get("revenues", [])
        expenses = context.get("expenses", [])
    
        max_len = max(len(revenues), len(expenses))
    
        for i in range(max_len):
            revenue = revenues[i] if i < len(revenues) else None
            expense = expenses[i] if i < len(expenses) else None
    
            row = []
    
            # Revenue Columns
            if revenue:
                row.append(revenue.date.strftime("%Y-%m-%d"))
                row.append(revenue.category.category)
                row.append("{:,.2f}".format(revenue.amount))
                row.append(revenue.description or "")
            else:
                row.extend(["", "", "", ""])  # Empty values if no revenue
    
            row.extend([""])  # Spacer Columns
    
            # Expense Columns
            if expense:
                row.append(expense.date.strftime("%Y-%m-%d"))
                row.append(expense.category.category)
                row.append("{:,.2f}".format(expense.amount))
                row.append(expense.description or "")
            else:
                row.extend(["", "", "", ""])  # Empty values if no expense
    
            ws.append(row)
    
        # Adjust column widths
        col_widths = [15, 20, 12, 25, 2, 15, 20, 12, 25]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    
        # Save to BytesIO
        excel_stream = BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)
    
        return self.prepare_export_response(excel_stream.getvalue(), "xlsx", self.model._meta.model_name)
    
class FundRevenueCreateView(PermissionMixin, CreateView):
    model = FundRevenue
    form_class = FundRevenueForm
    # fields = ['name', 'email', 'phone', 'type', 'category', 'join_date', 'balance']
    template_name = 'funds/fund_revenue_form.html'
    success_url = reverse_lazy('fund_list')
    permission_required = 'core.add_fundrevenue'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['fund'] = get_object_or_404(Fund, pk=self.kwargs['fund_pk'])
        return context

    def get_success_url(self):
        return reverse_lazy('fund_detail', kwargs={'pk': self.kwargs['fund_pk']})

    def form_valid(self, form):
        fund = get_object_or_404(Fund, pk=self.kwargs['fund_pk'])
        form.instance.fund = fund
        response = super().form_valid(form)
        messages.success(self.request, 'تم إضافة الايراد بنجاح.')
        return response

    def form_invalid(self, form):
        messages.error(self.request, 'حدث خطأ أثناء إضافة الايراد. يرجى المحاولة مرة أخرى.')
        return super().form_invalid(form)

class FundRevenueUpdateView(PermissionMixin, UpdateView):
    model = FundRevenue
    form_class = FundRevenueForm
    template_name = 'funds/fund_revenue_form.html'
    success_url = reverse_lazy('fund_list')
    permission_required = 'core.change_fundrevenue'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['fund'] = get_object_or_404(Fund, pk=self.kwargs['fund_pk'])
        return context

    def get_success_url(self):
        return reverse_lazy('fund_detail', kwargs={'pk': self.kwargs['fund_pk']})

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'تم تحديث الايراد بنجاح.')
        return response

    def form_invalid(self, form):
        messages.error(self.request, 'حدث خطأ أثناء تحديث الايراد. يرجى المحاولة مرة أخرى.')
        return super().form_invalid(form)

class FundExpenseCreateView(PermissionMixin, CreateView):
    model = FundExpense
    form_class = FundExpenseForm
    template_name = 'funds/fund_expense_form.html'
    success_url = reverse_lazy('fund_list')
    permission_required = 'core.add_fundexpense'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['fund'] = get_object_or_404(Fund, pk=self.kwargs['fund_pk'])
        return context

    def get_success_url(self):
        return reverse_lazy('fund_detail', kwargs={'pk': self.kwargs['fund_pk']})

    def form_valid(self, form):
        fund = get_object_or_404(Fund, pk=self.kwargs['fund_pk'])
        form.instance.fund = fund
        response = super().form_valid(form)
        messages.success(self.request, 'تم إضافة المصروف بنجاح.')
        return response

    def form_invalid(self, form):
        messages.error(self.request, 'حدث خطأ أثناء إضافة المصروف. يرجى المحاولة مرة أخرى.')
        return super().form_invalid(form)

class FundExpenseUpdateView(PermissionMixin, UpdateView):
    model = FundExpense
    form_class = FundExpenseForm
    template_name = 'funds/fund_expense_form.html'
    success_url = reverse_lazy('fund_list')
    permission_required = 'core.change_fundexpense'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['fund'] = get_object_or_404(Fund, pk=self.kwargs['fund_pk'])
        return context

    def get_success_url(self):
        return reverse_lazy('fund_detail', kwargs={'pk': self.kwargs['fund_pk']})

    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'تم تحديث المصروف بنجاح.')
        return response

    def form_invalid(self, form):
        messages.error(self.request, 'حدث خطأ أثناء تحديث المصروف. يرجى المحاولة مرة أخرى.')
        return super().form_invalid(form)

class FundRevenuesChangeView(PermissionMixin, View):
    template_name = 'funds/fund_revenues_change.html'
    permission_required = 'core.change_fundrevenue'

    def get(self, request, fund_pk):
        # Get the fund instance
        fund = get_object_or_404(Fund, pk=fund_pk)

        # Get the existing revenues for the fund
        revenues = FundRevenue.objects.filter(fund=fund)

        # Create a form instance for each revenue item
        revenue_forms = [
            FundRevenueForm(instance=revenue, prefix=f"revenue_{revenue.pk}")
            for revenue in revenues
        ]


        # Add a form to add a new revenue
        revenue_form = FundRevenueForm()

        context = {
            'fund': fund,
            'revenue_forms': revenue_forms,
            'revenue_form': revenue_form,
        }

        return render(request, self.template_name, context)

    def post(self, request, fund_pk):
        # Get the fund instance
        fund = get_object_or_404(Fund, pk=fund_pk)

        # Check if a delete action is triggered
        if "delete_revenue" in request.POST:
            revenue_id = request.POST.get("delete_revenue")
            revenue = get_object_or_404(FundRevenue, pk=revenue_id, fund=fund)
            revenue.delete()
            return redirect('fund_revenues_change', fund_pk=fund.pk)

        # Get the existing revenues for the fund
        revenues = FundRevenue.objects.filter(fund=fund)

        # Create form instances for each revenue item
        revenue_forms = [
            FundRevenueForm(
                request.POST,
                instance=revenue,
                prefix=f"revenue_{revenue.pk}"
            )
            for revenue in revenues
        ]

        # Create new revenue form to add new items
        revenue_form = FundRevenueForm(request.POST)

        # Update existing revenues
        for form in revenue_forms:
            if form.is_valid():
                form.save()

        # Add new revenue if valid
        if revenue_form.is_valid():
            new_revenue = revenue_form.save(commit=False)
            new_revenue.fund = fund
            new_revenue.save()

        # Redirect back to the same page with updated revenues
        return redirect('fund_revenues_change', fund_pk=fund.pk)

class FundExpensesChangeView(PermissionMixin, View):
    template_name = 'funds/fund_expenses_change.html'
    permission_required = 'core.change_fundexpense'

    def get(self, request, fund_pk):
        # Get the fund instance
        fund = get_object_or_404(Fund, pk=fund_pk)

        # Get the existing expenses for the fund
        expenses = FundExpense.objects.filter(fund=fund)

        # Create a form instance for each expense item with unique prefixes
        expense_forms = [
            FundExpenseForm(instance=expense, prefix=f"expense_{expense.pk}")
            for expense in expenses
        ]

        # Add a form to add a new expense
        expense_form = FundExpenseForm(prefix="new_expense")

        context = {
            'fund': fund,
            'expense_forms': expense_forms,
            'expense_form': expense_form,
        }

        return render(request, self.template_name, context)

    def post(self, request, fund_pk):
        # Get the fund instance
        fund = get_object_or_404(Fund, pk=fund_pk)

        # Check if a delete action is triggered
        if "delete_expense" in request.POST:
            expense_id = request.POST.get("delete_expense")
            expense = get_object_or_404(FundExpense, pk=expense_id, fund=fund)
            expense.delete()
            return redirect('fund_expenses_change', fund_pk=fund.pk)

        # Get the existing expenses for the fund
        expenses = FundExpense.objects.filter(fund=fund)

        # Create form instances for each expense item with unique prefixes
        expense_forms = [
            FundExpenseForm(
                request.POST,
                instance=expense,
                prefix=f"expense_{expense.pk}"
            )
            for expense in expenses
        ]

        # Create a new expense form to add new items
        expense_form = FundExpenseForm(request.POST, prefix="new_expense")

        # Flag to check if all forms are valid
        all_valid = all(form.is_valid() for form in expense_forms)

        # Update existing expenses if all forms are valid
        if all_valid:
            for form in expense_forms:
                form.save()

        # Add a new expense if valid
        if expense_form.is_valid():
            new_expense = expense_form.save(commit=False)
            new_expense.fund = fund
            new_expense.save()

        # Redirect back to the same page with updated expenses
        return redirect('fund_expenses_change', fund_pk=fund.pk)

class FundDeleteView(PermissionMixin, DeleteView):
    model = Fund
    template_name = 'fund_confirm_delete.html'
    success_url = reverse_lazy('fund_list')
    permission_required = 'core.delete_fund'
